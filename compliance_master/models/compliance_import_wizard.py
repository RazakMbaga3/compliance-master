import base64
import io
import re
import logging
from datetime import date, datetime

from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# ── Status normalisation map ───────────────────────────────────────────────────
EXCEL_STATUS_MAP = {
    'active': 'active', 'valid': 'active', 'available': 'active',
    '√': 'active', 'renewed': 'active', 'submitted yearly': 'active',
    'submitted quarterly': 'active', '90 days': 'active', 'yes': 'active',
    'under renewal': 'under_renewal', 'under renewal fees paid': 'under_renewal',
    'on process': 'under_renewal', 'under renewal fees': 'under_renewal',
    'not valid': 'overdue', 'expired': 'overdue',
    'inactive': 'inactive', 'not applied': 'inactive',
    'due': 'due',
}

# ── Compliance type inference keywords ────────────────────────────────────────
FLEET_AGENCIES = {
    'alliance insurance', 'land transport regulatory', 'weights & measure',
    'weights and measure', 'tanzania revenue authority',
}
PERIODIC_KEYWORDS = {
    'vat', 'paye', 'nssf', 'nhif', 'wcf', 'heslb', 'tuico',
    'withholding tax', 'witholding tax', 'city service levy',
    'skills & development levy', 'royalties', 'annual rent',
    'land rent', 'property tax', 'corporate tax', 'annual returns',
    'z report', 'audit',
}
CERT_KEYWORDS = {
    'certificate', 'calibration', 'correctness', 'registration',
    'incorporation', 'tin', 'vrn', 'iso',
}


def _map_status(raw):
    if not raw:
        return 'active'
    return EXCEL_STATUS_MAP.get(str(raw).strip().lower(), 'active')


def _infer_type(description, agency, frequency):
    """Guess compliance_type from description / agency / frequency text."""
    desc_l  = (description or '').lower()
    agency_l = (agency or '').lower()
    freq_l  = (frequency or '').lower()

    # Vehicle registration number pattern in description (e.g. T296DBZ)
    if re.match(r'^t\d{3}[a-z]{2,4}$', desc_l.strip()):
        return 'fleet'
    if any(a in agency_l for a in FLEET_AGENCIES):
        return 'fleet'

    if any(k in desc_l for k in PERIODIC_KEYWORDS):
        return 'periodic'
    if freq_l in ('monthly', 'quarterly') and not any(k in desc_l for k in CERT_KEYWORDS):
        return 'periodic'

    if any(k in desc_l for k in CERT_KEYWORDS):
        return 'certificate'

    return 'license'


def _infer_frequency(raw):
    """Map free-text frequency to Selection key."""
    if not raw:
        return False
    r = str(raw).strip().lower()
    if r in ('monthly', 'monthly '):
        return 'monthly'
    if r in ('quarterly', 'quarterly '):
        return 'quarterly'
    if 'bi' in r or 'half' in r or '2 year' in r or 'twice' in r:
        return 'bi_annual'
    if r in ('annual', 'yearly', 'once in a year', 'annually', '1', 'year'):
        return 'annual'
    if '15' in r:
        return '15_yrs'
    if '10' in r:
        return '10_yrs'
    if '5' in r:
        return '5_yrs'
    if 'lifetime' in r or 'project' in r or 'one time' in r or 'once' in r:
        return 'lifetime'
    return 'other'


def _parse_date(raw):
    """
    Robust parser for all date formats found in the Excel:
      25th jan2022  |  10thAug2021  |  2022-03-02  |  10.08.2023
      31.12.2022    |  01.04.2015   |  13/09/2022  |  datetime objects
      End of Project|  2030 (year)  |  01.07.2023
    Returns a date object or None.
    """
    if not raw:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.date() if isinstance(raw, datetime) else raw

    s = str(raw).strip()
    if not s or s.lower() in ('none', 'nan', 'end of project', 'n/a', '-'):
        return None

    # Pure year  e.g. "2030" or "2023"
    if re.fullmatch(r'\d{4}', s):
        return date(int(s), 12, 31)

    # ISO  2022-03-02
    m = re.fullmatch(r'(\d{4})-(\d{1,2})-(\d{1,2})(?:\s.*)?', s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # DD.MM.YYYY  or  DD/MM/YYYY  or  DD-MM-YYYY
    m = re.fullmatch(r'(\d{1,2})[./\-](\d{1,2})[./\-](\d{2,4})', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d)
        except ValueError:
            pass

    # "25th jan2022"  "10thAug2021"  "29thMar2016"
    m = re.match(
        r'(\d{1,2})(?:st|nd|rd|th)?\s*([a-zA-Z]+)\s*(\d{4})', s, re.IGNORECASE
    )
    if m:
        months = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        }
        mo_key = m.group(2).lower()[:3]
        mo = months.get(mo_key)
        if mo:
            try:
                return date(int(m.group(3)), mo, int(m.group(1)))
            except ValueError:
                pass

    # "01.07.2023" already covered above, but try strptime as fallback
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d',
                '%d %b %Y', '%d %B %Y', '%b %d, %Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    _logger.warning('compliance_import: could not parse date "%s"', s)
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  Wizard model
# ══════════════════════════════════════════════════════════════════════════════

class ComplianceImportWizard(models.TransientModel):
    _name = 'compliance.import.wizard'
    _description = 'Import Compliance Records from Excel'

    # ── Step 1 fields ─────────────────────────────────────────────────────────
    state = fields.Selection([
        ('upload',  'Upload'),
        ('preview', 'Preview'),
        ('done',    'Done'),
    ], default='upload', string='Step')

    excel_file    = fields.Binary(string='Excel File (.xlsx)', required=True)
    excel_filename = fields.Char(string='Filename')

    sheet_name = fields.Selection(
        selection='_get_sheet_choices',
        string='Sheet to Import',
        default='MASTER (2)',
    )

    skip_existing   = fields.Boolean(
        string='Skip duplicate records',
        default=True,
        help='If checked, records whose description+agency already exist will be skipped.',
    )
    clear_existing  = fields.Boolean(
        string='Delete ALL existing records before import',
        default=False,
        help='WARNING: this permanently deletes all current compliance records.',
    )

    # ── Step 2 / preview fields ────────────────────────────────────────────────
    preview_html  = fields.Html(string='Preview', readonly=True)
    total_found   = fields.Integer(string='Records Found',    readonly=True)
    total_valid   = fields.Integer(string='Ready to Import',  readonly=True)
    total_skipped = fields.Integer(string='Will Be Skipped',  readonly=True)
    total_errors  = fields.Integer(string='Rows With Errors', readonly=True)

    # ── Step 3 / result fields ─────────────────────────────────────────────────
    import_log    = fields.Text(string='Import Log', readonly=True)
    imported_count = fields.Integer(string='Records Imported', readonly=True)
    skipped_count  = fields.Integer(string='Records Skipped',  readonly=True)
    error_count    = fields.Integer(string='Errors',           readonly=True)

    # ══════════════════════════════════════════════════════════════════════════
    # Dynamic sheet list
    # ══════════════════════════════════════════════════════════════════════════

    @api.model
    def _get_sheet_choices(self):
        return [
            ('MASTER (2)', 'MASTER (2) — Main Compliance List'),
            ('HO', 'HO — Head Office'),
            ('Land & Mining', 'Land & Mining'),
            ('Plant', 'Plant'),
            ('BRL', 'BRL'),
            ('UCL', 'UCL'),
        ]

    # ══════════════════════════════════════════════════════════════════════════
    # Step 1 → analyse
    # ══════════════════════════════════════════════════════════════════════════

    def action_analyse(self):
        """Parse the file, build a preview summary, move to preview step."""
        self.ensure_one()
        if not self.excel_file:
            raise UserError('Please upload an Excel file first.')

        rows, errors = self._parse_excel()
        if errors and not rows:
            raise UserError('\n'.join(errors))

        # Check for duplicates
        existing_keys = set()
        if self.skip_existing:
            existing = self.env['compliance.record'].search_read(
                [], ['name', 'agency']
            )
            existing_keys = {
                (r['name'].strip().lower(), r['agency'].strip().lower())
                for r in existing
            }

        valid, skipped, err_rows = [], [], []
        for row in rows:
            key = (
                (row.get('name') or '').strip().lower(),
                (row.get('agency') or '').strip().lower(),
            )
            if not row.get('name') or not row.get('agency'):
                err_rows.append(row)
            elif self.skip_existing and key in existing_keys:
                skipped.append(row)
            else:
                valid.append(row)

        self.total_found   = len(rows)
        self.total_valid   = len(valid)
        self.total_skipped = len(skipped)
        self.total_errors  = len(err_rows) + len(errors)

        # Build preview HTML table (first 50 valid rows)
        self.preview_html  = self._build_preview_html(valid[:50], errors, skipped[:10])
        self.state = 'preview'
        return self._reload()

    # ══════════════════════════════════════════════════════════════════════════
    # Step 2 → import
    # ══════════════════════════════════════════════════════════════════════════

    def action_import(self):
        """Create all compliance.record rows in the database."""
        self.ensure_one()

        if self.clear_existing:
            self.env['compliance.record'].search([]).unlink()
            _logger.warning('compliance_import: all existing records deleted by user request')

        rows, _ = self._parse_excel()

        existing_keys = set()
        if self.skip_existing and not self.clear_existing:
            existing = self.env['compliance.record'].search_read([], ['name', 'agency'])
            existing_keys = {
                (r['name'].strip().lower(), r['agency'].strip().lower())
                for r in existing
            }

        # Cache divisions and departments to avoid repeated searches
        div_cache  = {}
        dept_cache = {}

        imported = skipped = errors = 0
        log_lines = []

        for i, row in enumerate(rows, 1):
            name   = (row.get('name') or '').strip()
            agency = (row.get('agency') or '').strip()

            if not name or not agency:
                errors += 1
                log_lines.append(f'[ROW {i}] SKIP – missing description or agency: {row}')
                continue

            key = (name.lower(), agency.lower())
            if self.skip_existing and key in existing_keys:
                skipped += 1
                continue

            # Resolve / create division
            div_name = (row.get('division') or '').strip()
            division_id = False
            if div_name:
                if div_name not in div_cache:
                    div = self.env['compliance.division'].search(
                        [('name', 'ilike', div_name)], limit=1
                    )
                    if not div:
                        div = self.env['compliance.division'].create({'name': div_name})
                        log_lines.append(f'[DIV] Created division: {div_name}')
                    div_cache[div_name] = div.id
                division_id = div_cache[div_name]

            # Resolve department
            dept_name = (row.get('department') or '').strip()
            department_id = False
            if dept_name:
                if dept_name not in dept_cache:
                    dept = self.env['hr.department'].search(
                        [('name', 'ilike', dept_name)], limit=1
                    )
                    if not dept:
                        dept = self.env['hr.department'].create({'name': dept_name})
                        log_lines.append(f'[DEPT] Created department: {dept_name}')
                    dept_cache[dept_name] = dept.id
                department_id = dept_cache[dept_name]

            try:
                vals = {
                    'name':               name,
                    'agency':             agency,
                    'compliance_type':    row.get('compliance_type', 'license'),
                    'division_id':        division_id,
                    'department_id':      department_id,
                    'location_custodian': row.get('custodian') or False,
                    'frequency':          row.get('frequency') or False,
                    'frequency_other':    row.get('frequency_raw') if row.get('frequency') == 'other' else False,
                    'origin_date':        row.get('origin_date') or False,
                    'valid_from':         row.get('valid_from') or False,
                    'expiry_date':        row.get('expiry_date') or False,
                    'renewal_date':       row.get('renewal_date') or False,
                    'remarks':            row.get('remarks') or False,
                    'vehicle_reg':        row.get('vehicle_reg') or False,
                    'notify_direct_days':  int(row.get('notify_direct', 0) or 0),
                    'notify_manager_days': int(row.get('notify_manager', 0) or 0),
                    'notify_head_days':    int(row.get('notify_head', 0) or 0),
                }

                rec = self.env['compliance.record'].create(vals)

                # Set status — write after create so state compute runs first
                raw_state = _map_status(row.get('status'))
                if raw_state in ('inactive', 'under_renewal'):
                    rec.write({'state': raw_state})

                imported += 1
                existing_keys.add(key)

            except Exception as e:
                errors += 1
                log_lines.append(f'[ROW {i}] ERROR creating "{name}": {e}')
                _logger.error('compliance_import row %s: %s', i, e)

        summary = (
            f'Import complete.\n'
            f'  Imported : {imported}\n'
            f'  Skipped  : {skipped}\n'
            f'  Errors   : {errors}\n'
        )
        if log_lines:
            summary += '\nDetail log:\n' + '\n'.join(log_lines)

        self.write({
            'state':          'done',
            'import_log':     summary,
            'imported_count': imported,
            'skipped_count':  skipped,
            'error_count':    errors,
        })
        return self._reload()

    # ══════════════════════════════════════════════════════════════════════════
    # Step back
    # ══════════════════════════════════════════════════════════════════════════

    def action_back(self):
        self.state = 'upload'
        return self._reload()

    def action_open_records(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Compliance Records',
            'res_model': 'compliance.record',
            'view_mode': 'tree,form',
        }

    # ══════════════════════════════════════════════════════════════════════════
    # Core Excel parser
    # ══════════════════════════════════════════════════════════════════════════

    def _parse_excel(self):
        """
        Parse the uploaded Excel workbook.
        Returns (list[dict], list[str_errors]).
        Supports both the MASTER (2) format and the simpler HO/Plant sheet format.
        """
        try:
            import openpyxl
        except ImportError:
            raise UserError(
                'The openpyxl library is required for Excel import.\n'
                'Install it with: pip install openpyxl'
            )

        file_data = base64.b64decode(self.excel_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)

        sheet = self.sheet_name or 'MASTER (2)'
        if sheet not in wb.sheetnames:
            raise UserError(
                f'Sheet "{sheet}" not found in the workbook.\n'
                f'Available sheets: {", ".join(wb.sheetnames)}'
            )

        ws = wb[sheet]

        # Detect which format this sheet uses
        if sheet == 'MASTER (2)':
            return self._parse_master_sheet(ws)
        else:
            return self._parse_simple_sheet(ws)

    # ── MASTER (2) parser ────────────────────────────────────────────────────
    #  Row 3 (idx 2): header row 1
    #  Row 4 (idx 3): header row 2 (sub-headers)
    #  Row 5+ (idx 4+): data
    #
    #  Col indices (0-based):
    #   0=Sr No  1=Division  2=Department  3=Agency  4=Description
    #   5=Custodian  6=Frequency  7=Resp Direct  8=Resp Manager  9=Resp Head
    #   10=Origin  11=Valid From  12=Valid To  13=Renewal
    #   14=Doc Original  15=Doc Current  16=Status
    #   17=Notify Direct  18=Notify Manager  19=Notify Head  20=Remarks

    def _parse_master_sheet(self, ws):
        rows_data = []
        errors    = []
        current_division   = ''
        current_department = ''
        skipped_header_rows = 4  # rows 1-4 are title/header

        all_rows = list(ws.iter_rows(min_row=skipped_header_rows + 1, values_only=True))

        for row_idx, row in enumerate(all_rows, start=skipped_header_rows + 1):
            # Extend short rows to avoid index errors
            row = list(row) + [None] * 25

            sr     = row[0]
            div    = self._str(row[1])
            dept   = self._str(row[2])
            agency = self._str(row[3])
            desc   = self._str(row[4])

            # Track running division / department from section headers
            if div:
                current_division = div
            if dept:
                current_department = dept

            # Skip blank / section-header rows
            if not desc or not agency:
                continue
            # Skip rows where description looks like a section title
            # (all caps, no agency makes sense — already handled above)

            # Parse notification days — may be blank or contain non-numeric
            def _days(val):
                try:
                    return int(float(str(val).strip())) if val else 0
                except (ValueError, TypeError):
                    return 0

            freq_raw = self._str(row[6])
            frequency = _infer_frequency(freq_raw)

            ctype = _infer_type(desc, agency, freq_raw)

            # Vehicle reg: if description matches a plate pattern use it
            vehicle_reg = False
            if re.match(r'^[Tt]\d{3}[A-Za-z]{2,4}$', desc.strip()):
                vehicle_reg = desc.strip().upper()

            rows_data.append({
                'division':        current_division,
                'department':      current_department,
                'agency':          agency,
                'name':            desc,
                'custodian':       self._str(row[5]),
                'frequency':       frequency,
                'frequency_raw':   freq_raw,
                'resp_direct':     self._str(row[7]),
                'resp_manager':    self._str(row[8]),
                'resp_head':       self._str(row[9]),
                'origin_date':     _parse_date(row[10]),
                'valid_from':      _parse_date(row[11]),
                'expiry_date':     _parse_date(row[12]),
                'renewal_date':    _parse_date(row[13]),
                'status':          self._str(row[16]),
                'notify_direct':   _days(row[17]),
                'notify_manager':  _days(row[18]),
                'notify_head':     _days(row[19]),
                'remarks':         self._str(row[20]),
                'compliance_type': ctype,
                'vehicle_reg':     vehicle_reg,
                '_row':            row_idx,
            })

        return rows_data, errors

    # ── Simple sheet parser (HO / Plant / BRL / UCL / Land & Mining) ─────────
    #  Row 3: header  |  Col: # | Description | Agency | Number | FROM | TO | Remarks | Status | Original | Department
    def _parse_simple_sheet(self, ws):
        rows_data = []
        errors    = []
        all_rows = list(ws.iter_rows(min_row=5, values_only=True))

        for row_idx, row in enumerate(all_rows, start=5):
            row = list(row) + [None] * 15
            desc   = self._str(row[1])
            agency = self._str(row[2])
            if not desc or not agency:
                continue

            rows_data.append({
                'division':        '',
                'department':      self._str(row[9]),
                'agency':          agency,
                'name':            desc,
                'custodian':       '',
                'frequency':       'annual',
                'frequency_raw':   '',
                'resp_direct':     '',
                'resp_manager':    '',
                'resp_head':       '',
                'origin_date':     None,
                'valid_from':      _parse_date(row[4]),
                'expiry_date':     _parse_date(row[5]),
                'renewal_date':    None,
                'status':          self._str(row[7]),
                'notify_direct':   30,
                'notify_manager':  20,
                'notify_head':     20,
                'remarks':         self._str(row[6]),
                'compliance_type': _infer_type(desc, agency, ''),
                'vehicle_reg':     self._str(row[3]) or False,
                '_row':            row_idx,
            })

        return rows_data, errors

    # ══════════════════════════════════════════════════════════════════════════
    # Helpers
    # ══════════════════════════════════════════════════════════════════════════

    @staticmethod
    def _str(val):
        if val is None:
            return ''
        s = str(val).strip()
        return '' if s.lower() in ('none', 'nan', 'false') else s

    def _build_preview_html(self, valid_rows, errors, skipped_rows):
        lines = [
            '<div style="font-family:Arial,sans-serif;font-size:12px;">',
            f'<p><strong>Records ready to import: {self.total_valid}</strong> &nbsp;|&nbsp; '
            f'Skipped (duplicates): {self.total_skipped} &nbsp;|&nbsp; '
            f'Errors: {self.total_errors}</p>',
        ]

        if errors:
            lines.append('<h4 style="color:#c0392b;">Parse Warnings</h4><ul>')
            for e in errors[:10]:
                lines.append(f'<li style="color:#c0392b;">{e}</li>')
            lines.append('</ul>')

        # Preview table
        lines.append(
            '<table border="1" cellpadding="4" cellspacing="0" '
            'style="border-collapse:collapse;width:100%;font-size:11px;">'
            '<thead style="background:#1a4f8a;color:#fff;">'
            '<tr><th>#</th><th>Type</th><th>Division</th><th>Description</th>'
            '<th>Agency</th><th>Expiry</th><th>Status</th><th>Notify Days</th></tr>'
            '</thead><tbody>'
        )
        STATE_COLOURS = {
            'active': '#27ae60', 'due': '#e67e22',
            'overdue': '#c0392b', 'under_renewal': '#2980b9', 'inactive': '#95a5a6',
        }
        STATE_LABELS = {
            'active': 'Active', 'due': 'Due', 'overdue': 'Overdue',
            'under_renewal': 'Under Renewal', 'inactive': 'Inactive',
        }
        for i, r in enumerate(valid_rows, 1):
            st = _map_status(r.get('status'))
            colour = STATE_COLOURS.get(st, '#555')
            exp = str(r.get('expiry_date') or '—')
            nd  = '/'.join(filter(None, [
                str(r['notify_direct'])  if r.get('notify_direct')  else '',
                str(r['notify_manager']) if r.get('notify_manager') else '',
                str(r['notify_head'])    if r.get('notify_head')    else '',
            ])) or '—'
            bg = '#fff8f8' if st == 'overdue' else ('#fffbf0' if st == 'due' else '#fff')
            lines.append(
                f'<tr style="background:{bg};">'
                f'<td>{i}</td>'
                f'<td>{r.get("compliance_type","")}</td>'
                f'<td>{r.get("division","")}</td>'
                f'<td>{(r.get("name") or "")[:60]}</td>'
                f'<td>{r.get("agency","")}</td>'
                f'<td>{exp}</td>'
                f'<td><span style="color:{colour};font-weight:bold;">'
                f'{STATE_LABELS.get(st, st)}</span></td>'
                f'<td>{nd}</td>'
                f'</tr>'
            )
        if self.total_valid > 50:
            lines.append(
                f'<tr><td colspan="8" style="text-align:center;color:#555;">'
                f'… and {self.total_valid - 50} more records (showing first 50)</td></tr>'
            )
        lines.append('</tbody></table>')

        if skipped_rows:
            lines.append(
                '<h4 style="color:#7f8c8d;margin-top:12px;">'
                'Sample duplicates that will be skipped:</h4><ul>'
            )
            for r in skipped_rows:
                lines.append(f'<li style="color:#7f8c8d;">{r.get("name")} — {r.get("agency")}</li>')
            lines.append('</ul>')

        lines.append('</div>')
        return ''.join(lines)

    def _reload(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
